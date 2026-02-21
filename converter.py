#!/usr/bin/env python3
"""
HKL JSON to XLSX Converter

Converts JSON order files to XLSX files using product-specific templates.
Templates preserve Excel formulas (columns N-W) and dictionary tables (columns X+).
Only data in columns A-L is replaced from JSON.

Usage:
    python3 converter.py setup                # Extract templates from existing xlsx files
    python3 converter.py convert file.json    # Convert one JSON file
    python3 converter.py convert-all          # Convert all JSON files in input dir

Steps:
    1. Run 'setup' once to extract templates from existing xlsx examples
    2. Run 'convert' or 'convert-all' to convert JSON files
"""

import json
import os
import re
import sys
import urllib.request
import urllib.error
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)

GPT_REVIEW_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")


# ── Configuration ──────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
INPUT_DIR = os.path.join(BASE_DIR, "importyzefordoprod")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
MAPPINGS_DIR = os.path.join(BASE_DIR, "mappings")
TRAINING_DIR = os.path.join(BASE_DIR, "10")

# Order-level JSON fields → row numbers in columns A-B
HEADER_ROWS = {
    'address': 2,
    'city': 3,
    'client': 4,
    'comment': 5,
    'commission': 6,
    'country': 7,
    'email': 8,
    'name': 9,
    'orderno': 10,
    'organizationIdent': 11,
    'phone': 12,
    'sentDate': 13,
    'tax': 14,
    'userIdent': 15,
    'zip': 16,
    'orderid': 17,
}

# Item-level JSON fields → row numbers in columns C-D
ITEM_ROWS = {
    'product': 2,
    'comment': 3,
    'orderpos': 4,
    'posid': 5,
    'product_description': 6,
    'commission': 7,
    'department': 8,
}

# Fields that should be written as numbers in column D
NUMERIC_ITEM_FIELDS = {'product', 'orderpos', 'posid'}


# ── Helper functions ───────────────────────────────────────────────────────────

def to_excel(val):
    """Convert a Python/JSON value to Excel cell value.
    None and empty strings become '<NULL>'.
    Booleans become 1.0 / 0.0.
    Numbers and non-empty strings pass through.
    """
    if val is None:
        return '<NULL>'
    if isinstance(val, bool):
        return 1.0 if val else 0.0
    if isinstance(val, str) and val.strip() == '':
        return '<NULL>'
    return val


def read_param_rows(ws):
    """Read column E of the template to get parameter_name → row mapping.
    Returns two dicts:
      - main_params: parameters without dots (e.g., WYMIAROWANIE_SLOPOW → row)
      - sub_params:  sub-field parameters with dots (e.g., WYMIAROWANIE_SLOPOW.TYP → row)
    """
    main_params = {}
    sub_params = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=5).value  # Column E
        if name and isinstance(name, str):
            if '.' in name:
                sub_params[name] = row
            else:
                main_params[name] = row
    return main_params, sub_params


# ── GPT API fallback ──────────────────────────────────────────────────────────

def get_api_key():
    """Load OpenAI API key from .env file or environment variable."""
    env_path = os.path.join(BASE_DIR, '.env')
    if os.path.exists(env_path):
        with open(env_path, 'r') as f:
            for line in f:
                line = line.strip()
                if line.startswith('OPENAI_API_KEY='):
                    key = line.split('=', 1)[1].strip().strip('"').strip("'")
                    if key and not key.startswith('sk-proj-PASTE'):
                        return key
    return os.environ.get('OPENAI_API_KEY')


def extract_base_param_names(parameters):
    """Extract base parameter names from JSON, excluding metadata suffixes."""
    suffixes = ('_ALIAS___DESCRIPTION', '___DESCRIPTION', '___TITLE',
                '___VISIBLE', '___DICT', '_ALIAS')
    base_names = []
    seen = set()
    for key in parameters:
        is_suffix = False
        for s in suffixes:
            if key.endswith(s):
                is_suffix = True
                break
        if not is_suffix and key not in seen:
            base_names.append(key)
            seen.add(key)
    return base_names


def build_gpt_prompt(item):
    """Build a GPT prompt to determine parameter ordering for a new product."""
    params = item.get('parameters', {})
    base_names = extract_base_param_names(params)

    prompt = (
        "You are helping create an Excel template for an HKL order system.\n"
        "I need the parameters ordered logically for a product spreadsheet.\n"
        "Parameters go in column E starting at row 2.\n\n"
        "Example ordering for VERTIKAL #14:\n"
        "ILOSC, KONFIGURACJA, RODZAJ, MODEL, KOLOR_SYSTEMU, KOLOR, PASEK, "
        "SZEROKOSC, WYS_RODZ, WYSOKOSC, WYMIAROWANIE_SLOPOW, ILOSC_PASK, "
        "KORALIK_OBSL, DLUGOSC_STER, MOTOR, ZASILANIE, PILOT, AUTOMATYKA, "
        "KIESZEN, KORALIK_DOL, MONTAZ, OCHRONA_CHILD_SAFETY, WYSOKOSC_MONTAZU, "
        "POW, INSTRUKCJA, FILM, CENA, CENA_SUMA, CENA_RABAT, DOPLATA, "
        "SUMA_BRUTTO, WARTOSC_KONCOWA, CENA_KONCOWA, OPIS_POZYCJI, OPIS_CENY, "
        "OPIS_RABATU\n\n"
        "The general pattern:\n"
        "1. ILOSC (quantity) first\n"
        "2. Product type/model params (MODEL, PROD, etc.)\n"
        "3. Color params (KOLOR, KOLOR_RAMY, KOLOR_SYSTE, etc.)\n"
        "4. Dimension params (SZEROKOSC, WYSOKOSC)\n"
        "5. Technical/feature params (STEROWANIE, etc.)\n"
        "6. Pricing: CENA, CENA_SUMA, CENA_RABAT, DOPLATA, SUMA_BRUTTO, "
        "WARTOSC_KONCOWA, CENA_KONCOWA\n"
        "7. Descriptions last: OPIS_POZYCJI, OPIS_CENY, OPIS_RABATU\n\n"
        f"New product: {item.get('department', '?')} (#{item.get('product', '?')})\n"
        f"Parameters: {', '.join(base_names)}\n\n"
        "Return ONLY a JSON array with these exact parameter names in the "
        "correct order. No explanation, just the JSON array."
    )
    return prompt


def call_gpt_api(prompt, api_key, model="gpt-4o-mini"):
    """Call OpenAI API using urllib (no extra dependencies)."""
    url = "https://api.openai.com/v1/chat/completions"

    payload = json.dumps({
        "model": model,
        "messages": [
            {"role": "system", "content": "You return only valid JSON."},
            {"role": "user", "content": prompt}
        ],
        "temperature": 0
    }).encode('utf-8')

    req = urllib.request.Request(url, data=payload, method='POST')
    req.add_header('Content-Type', 'application/json')
    req.add_header('Authorization', f'Bearer {api_key}')

    try:
        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))
            content = result['choices'][0]['message']['content'].strip()
            # Strip markdown code fences if present
            if content.startswith('```'):
                content = content.split('\n', 1)[1]
                content = content.rsplit('```', 1)[0].strip()
            return json.loads(content)
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8')
        print(f"  GPT API error ({e.code}): {error_body[:200]}")
        return None
    except Exception as e:
        print(f"  GPT API error: {e}")
        return None


def create_template_from_gpt(param_order, item):
    """Create a template xlsx from GPT-determined parameter ordering."""
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1 headers (matching existing templates)
    headers = {
        1: 'Nagłówek', 3: 'Pozycja', 5: 'Identyfikator', 6: 'Wart',
        7: 'Wart_opis', 8: 'Alias', 9: 'Alias_opis', 10: 'Tytuł',
        11: 'Widoczny', 12: 'Słownik', 14: 'Rej', 16: 'Poz',
    }
    for col, val in headers.items():
        ws.cell(row=1, column=col, value=val)

    # Column E: parameter names in GPT-determined order
    params = item.get('parameters', {})
    suffixes = ('_ALIAS___DESCRIPTION', '___DESCRIPTION', '___TITLE',
                '___VISIBLE', '___DICT', '_ALIAS')
    row = 2
    param_rows = {}

    for param_name in param_order:
        ws.cell(row=row, column=5, value=param_name)
        param_rows[param_name] = row
        row += 1

        # If the value is a dict, add sub-field rows
        val = params.get(param_name)
        if isinstance(val, dict):
            for sub_key in val:
                is_suffix = any(sub_key.endswith(s) for s in suffixes)
                if not is_suffix:
                    ws.cell(row=row, column=5, value=f"{param_name}.{sub_key}")
                    row += 1

    # ── Columns N-O: registration header formulas (universal) ──
    n_o = [
        (2, 'organizacja_kod',
         '=IF($B$11="Cozy","05",IF($B$11="HKL","04",'
         'IF($B$11="LuxanGmbH","03",IF($B$11="FENIX","02",'
         'IF($B$11="LuxanEwaKrawczyk","01","00")))))'),
        (3, 'organizacja', '=B11'),
        (4, 'uzytkownik', '=B15'),
        (5, 'ident', '=B15&"-"&B10'),
        (6, 'uwagi', '=$B$6'),
    ]
    for r, label, formula in n_o:
        ws.cell(row=r, column=14, value=label)
        ws.cell(row=r, column=15, value=formula)

    # ── Columns P-Q: position formulas ──
    # Universal formulas (reference only header/item rows in A-D)
    pq_universal = [
        (2,  'kom',                '="EO/"&$B$10&"/"&D4&"/"&$B$6&"-"&$D$7'),
        (3,  'odbiorca',           '=$B$15'),
        (4,  'dealer',             None),
        (5,  'kooperant',          None),
        (6,  'sprzedawca',         '=B11'),
        (7,  'adres_ident',        '=$B$15'),
        (8,  'adres_nazwa',        None),
        (9,  'adres_adres',        '=$B$2'),
        (10, 'adres_kodpoczt',     '=$B$16'),
        (11, 'adres_miejscowosc',  '=$B$3'),
        (12, 'adres_kraj',         '=$B$7'),
        (13, 'adres_nrtel',        '=$B$12'),
        (14, 'adres_email',        '=$B$8'),
        (15, 'adres_region',       None),
        (16, 'uwagi',              '=$D$3'),
    ]
    for r, label, formula in pq_universal:
        ws.cell(row=r, column=16, value=label)
        if formula:
            ws.cell(row=r, column=17, value=formula)

    # Dynamic formulas (adjusted to actual parameter row positions)
    ilosc_r = param_rows.get('ILOSC')
    cena_r = param_rows.get('CENA')
    cena_konc_r = param_rows.get('CENA_KONCOWA') or param_rows.get('WARTOSC_KONCOWA')
    opis_poz_r = param_rows.get('OPIS_POZYCJI')
    opis_ceny_r = param_rows.get('OPIS_CENY')
    opis_rab_r = param_rows.get('OPIS_RABATU')

    ws.cell(row=17, column=16, value='liczba')
    if ilosc_r:
        ws.cell(row=17, column=17, value=f'=F{ilosc_r}')

    ws.cell(row=18, column=16, value='ilosc')
    ws.cell(row=18, column=17, value=1)

    ws.cell(row=19, column=16, value='cena_podst')
    if cena_r:
        ws.cell(row=19, column=17,
                value=f'=IF(ISERROR($F${cena_r}/1),0,$F${cena_r})')

    ws.cell(row=20, column=16, value='rabat_proc')
    ws.cell(row=20, column=17,
            value='=ROUND(IF(ISERROR($Q$21/$Q$19),0,$Q$21/$Q$19),4)')

    ws.cell(row=21, column=16, value='rabat')
    ws.cell(row=21, column=17, value='=$Q$19-$Q$22')

    ws.cell(row=22, column=16, value='cena')
    if cena_konc_r:
        ws.cell(row=22, column=17,
                value=f'=IF(ISERROR($F${cena_konc_r}/1),Q19,$F${cena_konc_r})')
    else:
        ws.cell(row=22, column=17, value='=Q19')

    ws.cell(row=23, column=16, value='nazwa')
    if opis_poz_r:
        ws.cell(row=23, column=17,
                value=f'=$D$8&" "&$D$6&" "&$F${opis_poz_r}')
    else:
        ws.cell(row=23, column=17, value='=$D$8&" "&$D$6')

    ws.cell(row=24, column=16, value='cena_info')
    if opis_ceny_r:
        ws.cell(row=24, column=17, value=f'=$F${opis_ceny_r}')

    ws.cell(row=25, column=16, value='rabat_info')
    if opis_rab_r:
        ws.cell(row=25, column=17, value=f'=$F${opis_rab_r}')

    ws.cell(row=26, column=16, value='grupa_cenowa')
    ws.cell(row=26, column=17,
            value='=IF(ISERROR(FIND("#",$I$7,1)),"-",'
                  'MID($I$7,FIND("#",$I$7,1)+1,20))')

    return wb


# ── EFOR→PROD Parameter Translator ────────────────────────────────────────────

def parse_param_string(s):
    """Parse 'KEY=VAL, KEY=VAL, ...' string into dict."""
    if not s or not isinstance(s, str):
        return {}
    result = {}
    # Split on comma (with optional surrounding whitespace)
    for part in s.split(','):
        part = part.strip()
        if not part:
            continue
        if '=' in part:
            key, val = part.split('=', 1)
            result[key.strip()] = val.strip()
        # Skip parts without '='
    return result


def ingest_training_data(xlsx_path):
    """Read training xlsx, return dict: ASORTMENT → list of (input_dict, output_dict)."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        asortment = row[4]  # col E
        inp_str = row[5]    # col F = Dane wejściowe
        out_str = row[6]    # col G = Dane wyjściowe

        if not asortment or not inp_str or not out_str:
            continue

        inp_dict = parse_param_string(str(inp_str))
        out_dict = parse_param_string(str(out_str))

        if inp_dict and out_dict:
            groups[str(asortment).strip()].append((inp_dict, out_dict))

    wb.close()
    return dict(groups)


def learn_key_mapping(pairs):
    """Given list of (input_dict, output_dict) pairs, learn the mapping.

    For each output_key, find the best matching input_key by checking:
    1. Exact value match rate (copy)
    2. Value/10 match rate (divide10)
    3. Consistent value→value mapping (lookup/dictionary)

    Returns: {
        "key_map": {"PROD_KEY": {"source": "EFOR_KEY", "transform": "copy|divide10|lookup"}},
        "value_map": {"PROD_KEY": {"efor_val": "prod_val", ...}},
        "constants": {"PROD_KEY": "fixed_value"}
    }
    """
    n = len(pairs)
    if n == 0:
        return {"key_map": {}, "value_map": {}, "constants": {}}

    # Collect all output keys and their values
    all_output_keys = {}
    for inp, out in pairs:
        for k, v in out.items():
            if k not in all_output_keys:
                all_output_keys[k] = []
            all_output_keys[k].append(v)

    # Collect all input keys
    all_input_keys = set()
    for inp, out in pairs:
        all_input_keys.update(inp.keys())

    key_map = {}
    value_map = {}
    constants = {}

    # Preserve output key order from training data
    output_key_order = []
    seen_out = set()
    for _, out in pairs:
        for k in out:
            if k not in seen_out:
                output_key_order.append(k)
                seen_out.add(k)

    for out_key in output_key_order:
        out_vals = all_output_keys[out_key]

        # Check if constant (all same value)
        unique_vals = set(out_vals)
        if len(unique_vals) == 1:
            constants[out_key] = out_vals[0]
            continue

        best_key = None
        best_score = 0
        best_transform = 'copy'
        best_lookup = None

        for inp_key in all_input_keys:
            copy_score = 0
            div10_score = 0
            lookup = {}
            lookup_consistent = True

            for inp, out in pairs:
                inp_val = inp.get(inp_key, '')
                out_val = out.get(out_key, '')

                if not out_val or out_val == '-':
                    continue

                # Copy match
                if str(inp_val) == str(out_val):
                    copy_score += 1

                # Divide by 10 match
                try:
                    iv = float(inp_val)
                    ov = float(out_val)
                    if abs(iv / 10.0 - ov) < 0.01:
                        div10_score += 1
                except (ValueError, TypeError, ZeroDivisionError):
                    pass

                # Lookup consistency
                if inp_val in lookup:
                    if lookup[inp_val] != out_val:
                        lookup_consistent = False
                else:
                    lookup[inp_val] = out_val

            # Score evaluation - use fraction of non-empty pairs
            non_empty = sum(1 for _, out in pairs if out.get(out_key, '') not in ('', '-'))
            if non_empty == 0:
                continue

            # Pick best transform for this input key
            if copy_score >= div10_score:
                score = copy_score
                transform = 'copy'
            else:
                score = div10_score
                transform = 'divide10'

            # Lookup can be better if consistent and covers many pairs
            if lookup_consistent and len(lookup) > 0:
                lookup_score = sum(1 for inp, out in pairs
                                   if inp.get(inp_key, '') in lookup
                                   and lookup[inp.get(inp_key, '')] == out.get(out_key, ''))
                if lookup_score > score:
                    score = lookup_score
                    transform = 'lookup'

            if score > best_score:
                best_score = score
                best_key = inp_key
                best_transform = transform
                if transform == 'lookup':
                    best_lookup = lookup

        # Accept if best score > 60% of pairs
        threshold = n * 0.6
        if best_key and best_score > threshold:
            key_map[out_key] = {"source": best_key, "transform": best_transform}
            if best_transform == 'lookup' and best_lookup:
                value_map[out_key] = best_lookup
        elif best_key and best_score > 0:
            # Lower confidence but still the best match
            key_map[out_key] = {"source": best_key, "transform": best_transform}
            if best_transform == 'lookup' and best_lookup:
                value_map[out_key] = best_lookup

    # Collect unmapped output keys (not in key_map or constants)
    unmapped = {}
    for out_key in output_key_order:
        if out_key not in key_map and out_key not in constants:
            sample_values = all_output_keys[out_key][:5]
            sample_inputs = []
            for inp, out in pairs[:5]:
                if out_key in out:
                    sample_inputs.append(inp)
            unmapped[out_key] = {
                "sample_values": sample_values,
                "sample_inputs": sample_inputs,
            }

    return {"key_map": key_map, "value_map": value_map, "constants": constants,
            "unmapped": unmapped}


def save_mapping(asortment, mapping):
    """Save mapping dict to mappings/{safe_filename}.json"""
    os.makedirs(MAPPINGS_DIR, exist_ok=True)
    safe_name = re.sub(r'[^\w\s-]', '_', asortment).strip()
    safe_name = re.sub(r'\s+', '_', safe_name)
    path = os.path.join(MAPPINGS_DIR, f"{safe_name}.json")
    with open(path, 'w', encoding='utf-8') as f:
        json.dump({"asortment": asortment, **mapping}, f, indent=2, ensure_ascii=False)
    return path


def load_mapping(asortment):
    """Load mapping for an ASORTMENT. Returns None if not found.
    Tries exact match first, then fuzzy match on filename."""
    if not os.path.isdir(MAPPINGS_DIR):
        return None

    # Build safe name for exact match
    safe_name = re.sub(r'[^\w\s-]', '_', asortment).strip()
    safe_name = re.sub(r'\s+', '_', safe_name)
    exact_path = os.path.join(MAPPINGS_DIR, f"{safe_name}.json")
    if os.path.exists(exact_path):
        with open(exact_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    # Fuzzy match: case-insensitive substring
    asort_lower = asortment.lower()
    for fname in os.listdir(MAPPINGS_DIR):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(MAPPINGS_DIR, fname)
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        stored_asort = data.get('asortment', '').lower()
        # Check both directions of substring match
        if asort_lower in stored_asort or stored_asort in asort_lower:
            return data
        # Also check the filename without extension
        fname_lower = fname.replace('.json', '').replace('_', ' ').lower()
        if asort_lower in fname_lower or fname_lower in asort_lower:
            return data

    return None


def translate_params(efor_params, mapping):
    """Translate EFOR parameter dict → PROD parameter dict using mapping.

    For each entry in mapping['key_map']:
      - Find source value in efor_params
      - Apply transform: copy, divide10, or lookup in value_map
      - Handle NULL/empty → '-'
    Returns: ({"PROD_KEY": "value", ...}, set_of_gpt_suggested_keys)
    """
    result = {}
    gpt_keys = set()
    key_map = mapping.get('key_map', {})
    value_map = mapping.get('value_map', {})
    constants = mapping.get('constants', {})
    gpt_suggestions = mapping.get('gpt_suggestions', {})

    # Add constants first
    for prod_key, val in constants.items():
        result[prod_key] = val

    # Apply key mappings
    for prod_key, info in key_map.items():
        source_key = info['source']
        transform = info.get('transform', 'copy')

        # Find value in efor_params (try exact, then case-insensitive)
        raw_val = efor_params.get(source_key)
        if raw_val is None:
            # Case-insensitive fallback
            for k, v in efor_params.items():
                if k.lower() == source_key.lower():
                    raw_val = v
                    break

        if raw_val is None or str(raw_val).strip() in ('', '<NULL>', '<NONE>', 'None'):
            result[prod_key] = '-'
            continue

        raw_str = str(raw_val).strip()

        if transform == 'copy':
            result[prod_key] = raw_val if not isinstance(raw_val, str) else raw_str
        elif transform == 'divide10':
            try:
                num = float(raw_str)
                divided = num / 10.0
                # Keep as int if whole number
                if divided == int(divided):
                    result[prod_key] = int(divided)
                else:
                    result[prod_key] = round(divided, 2)
            except (ValueError, TypeError):
                result[prod_key] = raw_str
        elif transform == 'lookup':
            lut = value_map.get(prod_key, {})
            result[prod_key] = lut.get(raw_str, raw_str)
        else:
            result[prod_key] = raw_str

    # Apply GPT suggestions
    for prod_key, info in gpt_suggestions.items():
        gpt_keys.add(prod_key)
        source_key = info.get('source', 'manual')
        transform = info.get('transform', 'manual')

        if transform == 'manual' or source_key == 'manual':
            result[prod_key] = '?'
            continue

        # Find value in efor_params
        raw_val = efor_params.get(source_key)
        if raw_val is None:
            for k, v in efor_params.items():
                if k.lower() == source_key.lower():
                    raw_val = v
                    break

        if raw_val is None or str(raw_val).strip() in ('', '<NULL>', '<NONE>', 'None'):
            result[prod_key] = '-'
            continue

        raw_str = str(raw_val).strip()

        if transform == 'copy':
            result[prod_key] = raw_val if not isinstance(raw_val, str) else raw_str
        elif transform == 'divide10':
            try:
                num = float(raw_str)
                divided = num / 10.0
                if divided == int(divided):
                    result[prod_key] = int(divided)
                else:
                    result[prod_key] = round(divided, 2)
            except (ValueError, TypeError):
                result[prod_key] = raw_str
        elif transform == 'lookup':
            lut = info.get('value_map', {})
            result[prod_key] = lut.get(raw_str, raw_str)
        else:
            result[prod_key] = raw_str

    return result, gpt_keys


def match_asortment(department, product_description=None):
    """Match a JSON department/product_description to a known ASORTMENT mapping.

    Strategy:
    1. Exact match on department
    2. Case-insensitive substring match
    3. Try product_description too
    """
    if not os.path.isdir(MAPPINGS_DIR):
        return None

    candidates = []
    for fname in os.listdir(MAPPINGS_DIR):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(MAPPINGS_DIR, fname)
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        candidates.append((data.get('asortment', ''), fpath, data))

    if not candidates:
        return None

    dept_lower = (department or '').lower().strip()

    # Exact match
    for asort, fpath, data in candidates:
        if asort.lower().strip() == dept_lower:
            return data

    # Substring match (both directions)
    for asort, fpath, data in candidates:
        asort_lower = asort.lower().strip()
        if dept_lower and (dept_lower in asort_lower or asort_lower in dept_lower):
            return data

    # Try with product_description
    if product_description:
        desc_lower = product_description.lower().strip()
        for asort, fpath, data in candidates:
            asort_lower = asort.lower().strip()
            if desc_lower in asort_lower or asort_lower in desc_lower:
                return data

    return None


def extract_flat_params(item):
    """Extract flat EFOR parameters from a JSON item's 'parameters' dict.

    Flattens nested dicts and strips metadata suffixes.
    Returns: {"KEY": "value", ...}
    """
    params = item.get('parameters', {})
    suffixes = ('_ALIAS___DESCRIPTION', '___DESCRIPTION', '___TITLE',
                '___VISIBLE', '___DICT', '_ALIAS')
    flat = {}
    for key, val in params.items():
        # Skip metadata keys
        if any(key.endswith(s) for s in suffixes):
            continue
        if isinstance(val, dict):
            # Flatten nested dict (e.g., WYMIAROWANIE_SLOPOW.TYP)
            for sub_key, sub_val in val.items():
                if not any(sub_key.endswith(s) for s in suffixes):
                    flat[f"{key}.{sub_key}"] = sub_val
        else:
            flat[key] = val
    return flat


def translate_json(json_path, output_path=None):
    """Translate a JSON EFOR order to PROD output xlsx.

    1. Read JSON, extract items
    2. For each item: match ASORTMENT, load mapping, translate params
    3. Save wynik.xlsx
    """
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    items = data.get('items', [])
    if not items:
        print(f"  No items in {json_path}")
        return

    base_name = os.path.splitext(os.path.basename(json_path))[0]
    if output_path is None:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(OUTPUT_DIR, f"wynik_{base_name}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wynik"

    headers_written = False
    current_row = 2  # row 1 for headers
    all_gpt_columns = set()

    for seq, item in enumerate(items, 1):
        department = item.get('department', '')
        product_desc = item.get('product_description', '')
        print(f"  [{seq}/{len(items)}] {department} / {product_desc}")

        # Try to load mapping
        mapping = match_asortment(department, product_desc)
        if mapping is None:
            mapping = load_mapping(department)

        if mapping is None:
            print(f"    No mapping found for '{department}'. Trying GPT fallback...")
            efor_params = extract_flat_params(item)
            mapping = gpt_translate_fallback(efor_params, department)
            if mapping is None:
                print(f"    SKIP: no mapping and GPT fallback failed.")
                continue

        efor_params = extract_flat_params(item)
        prod_params, gpt_keys = translate_params(efor_params, mapping)

        if not prod_params:
            print(f"    SKIP: translation produced no output parameters.")
            continue

        # Write headers on first successful translation
        if not headers_written:
            for col_idx, key in enumerate(prod_params.keys(), 1):
                ws.cell(row=1, column=col_idx, value=key)
                if key in gpt_keys:
                    all_gpt_columns.add(col_idx)
            headers_written = True

        # Write data row
        # Use header order from row 1 for consistency
        header_keys = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for col_idx, key in enumerate(header_keys, 1):
            val = prod_params.get(key, '-')
            ws.cell(row=current_row, column=col_idx, value=val)

        # If this item has keys not in header, append them
        for key in prod_params:
            if key not in header_keys:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col, value=key)
                ws.cell(row=current_row, column=new_col, value=prod_params[key])
                if key in gpt_keys:
                    all_gpt_columns.add(new_col)

        current_row += 1
        n_gpt = len(gpt_keys & set(prod_params.keys()))
        n_total = len(prod_params)
        if n_gpt:
            print(f"    -> {n_total} PROD parameters ({n_gpt} GPT-suggested, marked RED)")
        else:
            print(f"    -> {n_total} PROD parameters")

    # Apply red highlighting to GPT-suggested columns
    if all_gpt_columns:
        for col_idx in all_gpt_columns:
            # Header cell
            ws.cell(row=1, column=col_idx).fill = GPT_REVIEW_FILL
            # Data cells
            for r in range(2, current_row):
                ws.cell(row=r, column=col_idx).fill = GPT_REVIEW_FILL

    wb.save(output_path)
    wb.close()
    print(f"\n  Output saved: {output_path}")
    print(f"  {current_row - 2} items translated")


def gpt_translate_fallback(efor_params, asortment):
    """When no mapping exists, ask GPT to suggest key+value translations."""
    api_key = get_api_key()
    if not api_key:
        print("    No API key for GPT fallback (set OPENAI_API_KEY in .env)")
        return None

    # Build a sample of known mappings for context
    example = ""
    if os.path.isdir(MAPPINGS_DIR):
        for fname in os.listdir(MAPPINGS_DIR):
            if fname.endswith('.json'):
                fpath = os.path.join(MAPPINGS_DIR, fname)
                with open(fpath, 'r', encoding='utf-8') as f:
                    ex_data = json.load(f)
                km = ex_data.get('key_map', {})
                if len(km) > 5:
                    example_pairs = list(km.items())[:8]
                    example = (f"\nExample mapping for '{ex_data.get('asortment', '?')}':\n"
                               + "\n".join(f"  {pk} <- {info['source']} ({info['transform']})"
                                           for pk, info in example_pairs))
                    break

    param_list = "\n".join(f"  {k} = {v}" for k, v in list(efor_params.items())[:30])
    prompt = (
        f"You are helping translate EFOR order parameters to PROD manufacturing parameters.\n"
        f"Product type (ASORTMENT): {asortment}\n\n"
        f"EFOR parameters:\n{param_list}\n"
        f"{example}\n\n"
        f"For each EFOR parameter, suggest the PROD key name and transform type.\n"
        f"Return JSON with this structure:\n"
        f'{{"key_map": {{"PROD_KEY": {{"source": "EFOR_KEY", "transform": "copy|divide10|lookup"}}, ...}}, '
        f'"value_map": {{}}, "constants": {{}}}}\n'
        f"Common patterns: SZEROKOSC→B (divide10), WYSOKOSC→H (divide10), "
        f"KOLOR→KOLOR (copy), MODEL→MODEL (copy).\n"
        f"Return ONLY valid JSON."
    )

    result = call_gpt_api(prompt, api_key)
    if result and isinstance(result, dict) and 'key_map' in result:
        # Save as pending mapping for review
        result.setdefault('value_map', {})
        result.setdefault('constants', {})
        path = save_mapping(f"_gpt_{asortment}", result)
        print(f"    GPT mapping saved (pending review): {path}")
        return result

    return None


def gpt_suggest_unmapped(unmapped_info, existing_key_map, existing_constants, asortment):
    """Ask GPT to suggest mappings for unmapped PROD keys.

    Args:
        unmapped_info: dict of {PROD_KEY: {"sample_values": [...], "sample_inputs": [...]}}
        existing_key_map: already-mapped key_map dict
        existing_constants: already-mapped constants dict
        asortment: product type name

    Returns: {"PROD_KEY": {"source": ..., "transform": ..., "description_pl": ..., "confidence": ..., "reason": ...}, ...}
    """
    if not unmapped_info:
        return {}

    api_key = get_api_key()
    if not api_key:
        print("    No API key for GPT suggestions (set OPENAI_API_KEY in .env)")
        return {}

    # Build context about already-mapped keys
    mapped_summary = []
    for prod_key, info in existing_key_map.items():
        mapped_summary.append(f"  {prod_key} <- {info['source']} ({info.get('transform', 'copy')})")
    mapped_text = "\n".join(mapped_summary[:30]) if mapped_summary else "  (none)"

    const_summary = []
    for prod_key, val in existing_constants.items():
        const_summary.append(f"  {prod_key} = {val}")
    const_text = "\n".join(const_summary[:20]) if const_summary else "  (none)"

    # Build unmapped keys info
    unmapped_items = []
    for prod_key, info in unmapped_info.items():
        samples = info.get("sample_values", [])
        inp_keys = set()
        for inp_dict in info.get("sample_inputs", []):
            inp_keys.update(inp_dict.keys())
        inp_sample = {}
        if info.get("sample_inputs"):
            inp_sample = info["sample_inputs"][0]
        unmapped_items.append(
            f"  PROD key: {prod_key}\n"
            f"    Sample output values: {samples}\n"
            f"    Available input keys: {sorted(inp_keys)[:20]}\n"
            f"    Sample input row: {dict(list(inp_sample.items())[:10])}"
        )
    unmapped_text = "\n".join(unmapped_items)

    prompt = (
        f"You are helping map EFOR (input) parameters to PROD (output) parameters "
        f"for product type: {asortment}.\n\n"
        f"Already-mapped PROD keys (PROD <- EFOR source):\n{mapped_text}\n\n"
        f"Constants:\n{const_text}\n\n"
        f"The following PROD output keys could NOT be automatically matched to any "
        f"EFOR input key. For each one, analyze the key name and sample values to "
        f"suggest the best mapping.\n\n"
        f"Unmapped keys:\n{unmapped_text}\n\n"
        f"For each unmapped PROD key, return a JSON object with:\n"
        f'- "source": the EFOR input key name to map from (or "manual" if no match)\n'
        f'- "transform": "copy", "divide10", "lookup", or "manual"\n'
        f'- "description_pl": brief description in POLISH of what this parameter is\n'
        f'- "confidence": "high", "medium", or "low"\n'
        f'- "reason": brief English explanation of why you chose this mapping\n'
        f'- "value_map": optional dict mapping input values to output values (for lookup transform)\n\n'
        f"Guidelines:\n"
        f"- If an unmapped PROD key name is similar to an already-mapped key, name it similarly\n"
        f"- KOMPONENTDE might be a concatenation of component fields\n"
        f"- MODEL might be a truncated version of a system model field\n"
        f"- Keys ending in DE/EN might be German/English translations\n"
        f"- SYST_ prefix usually refers to system-level parameters\n"
        f"- If no reasonable match, use source='manual' and transform='manual'\n\n"
        f"Return ONLY a valid JSON object where keys are the PROD key names."
    )

    result = call_gpt_api(prompt, api_key)
    if result and isinstance(result, dict):
        # Validate structure
        clean = {}
        for prod_key, suggestion in result.items():
            if isinstance(suggestion, dict) and 'source' in suggestion:
                clean[prod_key] = {
                    "source": suggestion.get("source", "manual"),
                    "transform": suggestion.get("transform", "manual"),
                    "description_pl": suggestion.get("description_pl", ""),
                    "confidence": suggestion.get("confidence", "low"),
                    "reason": suggestion.get("reason", ""),
                }
                if suggestion.get("value_map"):
                    clean[prod_key]["value_map"] = suggestion["value_map"]
        return clean

    return {}


# ── Template setup ─────────────────────────────────────────────────────────────

def setup_templates():
    """Scan existing xlsx files and extract one template per product type."""
    os.makedirs(TEMPLATES_DIR, exist_ok=True)

    if not os.path.isdir(INPUT_DIR):
        print(f"Input directory not found: {INPUT_DIR}")
        return {}

    xlsx_files = sorted(f for f in os.listdir(INPUT_DIR) if f.endswith('.xlsx'))
    found = {}

    for fname in xlsx_files:
        fpath = os.path.join(INPUT_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath)
            ws = wb.active

            product_raw = ws.cell(row=2, column=4).value  # D2 = product ID
            department = ws.cell(row=8, column=4).value    # D8 = department

            if product_raw is not None:
                product_key = str(int(float(product_raw)))

                if product_key not in found:
                    import shutil
                    dest = os.path.join(TEMPLATES_DIR, f"template_{product_key}.xlsx")
                    wb.close()
                    shutil.copy2(fpath, dest)
                    found[product_key] = {'source': fname, 'department': department}
                    print(f"  Product {product_key:>3} ({department}) <- {fname}")
                    continue

            wb.close()
        except Exception as e:
            print(f"  Skip {fname}: {e}")

    print(f"\n{len(found)} templates extracted to {TEMPLATES_DIR}/")
    return found


# ── Data filling ───────────────────────────────────────────────────────────────

def fill_worksheet(ws, order, item):
    """Fill columns A-L of a template worksheet with JSON data."""

    # ── Columns A-B: order header ──
    for field, row in HEADER_ROWS.items():
        ws.cell(row=row, column=1, value=field)           # A
        ws.cell(row=row, column=2, value=to_excel(order.get(field)))  # B

    # ── Columns C-D: item metadata ──
    for field, row in ITEM_ROWS.items():
        val = item.get(field)
        ws.cell(row=row, column=3, value=field)            # C

        if field in NUMERIC_ITEM_FIELDS and val is not None:
            try:
                ws.cell(row=row, column=4, value=float(val))  # D as number
            except (ValueError, TypeError):
                ws.cell(row=row, column=4, value=to_excel(val))
        else:
            ws.cell(row=row, column=4, value=to_excel(val))  # D

    # ── Columns E-L: parameters ──
    params = item.get('parameters', {})
    main_param_rows, sub_param_rows = read_param_rows(ws)

    def write_param_row(row, param_name, val, desc, alias, alias_desc, title, visible, is_dict):
        """Write a single parameter row to columns E-L."""
        ws.cell(row=row, column=5,  value=param_name)               # E: name
        ws.cell(row=row, column=6,  value=to_excel(val))             # F: value
        ws.cell(row=row, column=7,  value=to_excel(desc))            # G: description
        ws.cell(row=row, column=8,  value=to_excel(alias))           # H: alias
        ws.cell(row=row, column=9,  value=to_excel(alias_desc))      # I: alias desc
        ws.cell(row=row, column=10, value=title or '')               # J: title
        ws.cell(row=row, column=11, value=1.0 if visible else 0.0)   # K: visible
        ws.cell(row=row, column=12, value=1.0 if is_dict else 0.0)   # L: dict

    for param_name, row in main_param_rows.items():
        if param_name not in params:
            continue

        val = params.get(param_name)

        # Handle nested dict values (e.g., WYMIAROWANIE_SLOPOW with sub-fields)
        if isinstance(val, dict):
            # Write '<NULL>' for the main parameter row
            desc   = params.get(f'{param_name}___DESCRIPTION')
            alias  = params.get(f'{param_name}_ALIAS')
            adesc  = params.get(f'{param_name}_ALIAS___DESCRIPTION')
            title  = params.get(f'{param_name}___TITLE')
            vis    = params.get(f'{param_name}___VISIBLE')
            isdict = params.get(f'{param_name}___DICT')
            write_param_row(row, param_name, '<NULL>', desc, alias, adesc, title, vis, isdict)

            # Fill sub-field rows (e.g., WYMIAROWANIE_SLOPOW.TYP)
            for sub_key, sub_row in sub_param_rows.items():
                if not sub_key.startswith(param_name + '.'):
                    continue
                field_name = sub_key.split('.', 1)[1]  # e.g., "TYP"
                sub_val    = val.get(field_name)
                sub_desc   = val.get(f'{field_name}___DESCRIPTION')
                sub_alias  = val.get(f'{field_name}_ALIAS')
                sub_adesc  = val.get(f'{field_name}_ALIAS___DESCRIPTION')
                sub_title  = val.get(f'{field_name}___TITLE')
                sub_vis    = val.get(f'{field_name}___VISIBLE')
                sub_isdict = val.get(f'{field_name}___DICT')
                write_param_row(sub_row, sub_key, sub_val, sub_desc,
                                sub_alias, sub_adesc, sub_title, sub_vis, sub_isdict)
        else:
            desc   = params.get(f'{param_name}___DESCRIPTION')
            alias  = params.get(f'{param_name}_ALIAS')
            adesc  = params.get(f'{param_name}_ALIAS___DESCRIPTION')
            title  = params.get(f'{param_name}___TITLE')
            vis    = params.get(f'{param_name}___VISIBLE')
            isdict = params.get(f'{param_name}___DICT')
            write_param_row(row, param_name, val, desc, alias, adesc, title, vis, isdict)


# ── Conversion ─────────────────────────────────────────────────────────────────

def convert_json(json_path, output_dir=None):
    """Convert a JSON file to XLSX files (one per item/position)."""
    if output_dir is None:
        output_dir = OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    base_name = os.path.splitext(os.path.basename(json_path))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    items = data.get('items', [])
    if not items:
        print(f"  No items in {json_path}")
        return []

    product_counters = {}
    created = []

    for seq, item in enumerate(items, 1):
        product_id = str(item.get('product', ''))
        department = item.get('department', '?')

        template_path = os.path.join(TEMPLATES_DIR, f"template_{product_id}.xlsx")
        if not os.path.exists(template_path):
            # GPT fallback: create template for unknown product type
            print(f"  No template for product {product_id} ({department}). Calling GPT API...")
            api_key = get_api_key()
            if not api_key:
                print(f"  SKIP item {seq}: no API key (set OPENAI_API_KEY in .env)")
                continue

            prompt = build_gpt_prompt(item)
            param_order = call_gpt_api(prompt, api_key)
            if not param_order:
                print(f"  SKIP item {seq}: GPT API call failed")
                continue

            os.makedirs(TEMPLATES_DIR, exist_ok=True)
            template_wb = create_template_from_gpt(param_order, item)
            template_wb.save(template_path)
            template_wb.close()
            print(f"  Template created: template_{product_id}.xlsx")
            print(f"  Note: formulas in columns N-W may need manual setup.")

        # Per-product-type counter for file naming
        product_counters[product_id] = product_counters.get(product_id, 0) + 1
        ppos = product_counters[product_id]

        # Load template (preserves formulas and dictionaries)
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Fill data
        fill_worksheet(ws, data, item)

        # Save
        out_name = f"{base_name}({timestamp})-{seq}({ppos}).xlsx"
        out_path = os.path.join(output_dir, out_name)
        wb.save(out_path)
        wb.close()

        created.append(out_name)
        print(f"  [{seq}/{len(items)}] {out_name}  (product {product_id}, {department})")

    print(f"  -> {len(created)}/{len(items)} files created")
    return created


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    usage = """
HKL JSON to XLSX Converter

Commands:
  python3 converter.py setup               Extract templates from existing xlsx files
  python3 converter.py convert <file.json>  Convert a single JSON file
  python3 converter.py convert-all          Convert all JSON files in input directory
  python3 converter.py list-templates       Show available templates

EFOR→PROD Translator:
  python3 converter.py learn [path.xlsx]    Learn mappings from training data
  python3 converter.py translate <file.json> [output.xlsx]  Translate EFOR→PROD
  python3 converter.py list-mappings        Show learned mappings with stats

Directories:
  Templates:  ./templates/
  Input:      ./importyzefordoprod/
  Output:     ./output/
  Mappings:   ./mappings/
  Training:   ./10/
"""

    if len(sys.argv) < 2:
        print(usage)
        sys.exit(1)

    cmd = sys.argv[1]

    if cmd == 'setup':
        print("Extracting templates from existing xlsx files...\n")
        setup_templates()

    elif cmd == 'convert':
        if len(sys.argv) < 3:
            print("Usage: python3 converter.py convert <file.json>")
            sys.exit(1)

        json_path = sys.argv[2]
        # Try as-is, then relative to INPUT_DIR
        if not os.path.exists(json_path):
            alt = os.path.join(INPUT_DIR, json_path)
            if os.path.exists(alt):
                json_path = alt
            else:
                print(f"File not found: {json_path}")
                sys.exit(1)

        if not os.path.isdir(TEMPLATES_DIR) or not os.listdir(TEMPLATES_DIR):
            print("No templates found. Run 'python3 converter.py setup' first.")
            sys.exit(1)

        print(f"Converting: {os.path.basename(json_path)}")
        convert_json(json_path)

    elif cmd == 'convert-all':
        if not os.path.isdir(TEMPLATES_DIR) or not os.listdir(TEMPLATES_DIR):
            print("No templates found. Run 'python3 converter.py setup' first.")
            sys.exit(1)

        json_files = sorted(
            os.path.join(INPUT_DIR, f)
            for f in os.listdir(INPUT_DIR)
            if f.endswith('.json')
        )

        if not json_files:
            print(f"No JSON files found in {INPUT_DIR}")
            sys.exit(1)

        print(f"Found {len(json_files)} JSON files\n")
        total = 0
        for jf in json_files:
            print(f"Converting: {os.path.basename(jf)}")
            created = convert_json(jf)
            total += len(created)
            print()

        print(f"Done. {total} xlsx files created in {OUTPUT_DIR}/")

    elif cmd == 'list-templates':
        if not os.path.isdir(TEMPLATES_DIR):
            print("No templates directory. Run 'python3 converter.py setup' first.")
            sys.exit(1)

        templates = sorted(f for f in os.listdir(TEMPLATES_DIR) if f.endswith('.xlsx'))
        if not templates:
            print("No templates found.")
        else:
            print("Available templates:")
            for t in templates:
                product_id = t.replace('template_', '').replace('.xlsx', '')
                try:
                    wb = openpyxl.load_workbook(os.path.join(TEMPLATES_DIR, t), read_only=True)
                    ws = wb.active
                    dept = ws.cell(row=8, column=4).value or '?'
                    desc = ws.cell(row=6, column=4).value or '?'
                    wb.close()
                    print(f"  Product {product_id:>3}: {dept} / {desc}")
                except Exception:
                    print(f"  Product {product_id:>3}: {t}")

    elif cmd == 'learn':
        # python3 converter.py learn [path_to_xlsx]
        if len(sys.argv) >= 3:
            xlsx_path = sys.argv[2]
        else:
            # Default: try 10/3.xlsx, then 10/10.xlsx
            xlsx_path = os.path.join(TRAINING_DIR, "3.xlsx")
            if not os.path.exists(xlsx_path):
                xlsx_path = os.path.join(TRAINING_DIR, "10.xlsx")

        if not os.path.exists(xlsx_path):
            print(f"Training file not found: {xlsx_path}")
            sys.exit(1)

        print(f"Learning from: {xlsx_path}\n")
        print("Reading training data...")
        groups = ingest_training_data(xlsx_path)
        print(f"  Found {len(groups)} ASORTMENT types, "
              f"{sum(len(v) for v in groups.values())} total pairs\n")

        for asort, pairs in sorted(groups.items()):
            print(f"  Learning: {asort} ({len(pairs)} examples)...", end=" ")
            mapping = learn_key_mapping(pairs)
            n_keys = len(mapping['key_map'])
            n_const = len(mapping['constants'])
            unmapped = mapping.pop('unmapped', {})
            n_unmapped = len(unmapped)
            print(f"-> {n_keys} key mappings, {n_const} constants, {n_unmapped} unmapped")

            if unmapped:
                print(f"    Unmapped keys: {', '.join(unmapped.keys())}")
                print(f"    Asking GPT for suggestions...", end=" ")
                suggestions = gpt_suggest_unmapped(
                    unmapped, mapping['key_map'], mapping['constants'], asort)
                if suggestions:
                    mapping['gpt_suggestions'] = suggestions
                    print(f"-> {len(suggestions)} GPT suggestions")
                    for pk, info in suggestions.items():
                        print(f"      {pk}: {info.get('source')} ({info.get('transform')}) "
                              f"[{info.get('confidence')}] - {info.get('description_pl', '')}")
                else:
                    print("-> no suggestions returned")

            save_mapping(asort, mapping)

        print(f"\nDone. Mappings saved to {MAPPINGS_DIR}/")

    elif cmd == 'translate':
        if len(sys.argv) < 3:
            print("Usage: python3 converter.py translate <file.json> [output.xlsx]")
            sys.exit(1)

        json_path = sys.argv[2]
        # Try as-is, then relative to INPUT_DIR
        if not os.path.exists(json_path):
            alt = os.path.join(INPUT_DIR, json_path)
            if os.path.exists(alt):
                json_path = alt
            else:
                print(f"File not found: {json_path}")
                sys.exit(1)

        output_path = sys.argv[3] if len(sys.argv) >= 4 else None

        if not os.path.isdir(MAPPINGS_DIR) or not os.listdir(MAPPINGS_DIR):
            print("No mappings found. Run 'python3 converter.py learn' first.")
            sys.exit(1)

        print(f"Translating: {os.path.basename(json_path)}\n")
        translate_json(json_path, output_path)

    elif cmd == 'list-mappings':
        if not os.path.isdir(MAPPINGS_DIR):
            print("No mappings directory. Run 'python3 converter.py learn' first.")
            sys.exit(1)

        files = sorted(f for f in os.listdir(MAPPINGS_DIR) if f.endswith('.json'))
        if not files:
            print("No mappings found.")
        else:
            print(f"Learned mappings ({len(files)}):\n")
            for fname in files:
                fpath = os.path.join(MAPPINGS_DIR, fname)
                with open(fpath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                asort = data.get('asortment', '?')
                n_keys = len(data.get('key_map', {}))
                n_vals = len(data.get('value_map', {}))
                n_const = len(data.get('constants', {}))
                n_gpt = len(data.get('gpt_suggestions', {}))
                print(f"  {asort}")
                print(f"    File: {fname}")
                print(f"    Keys: {n_keys} mappings, {n_vals} value lookups, "
                      f"{n_const} constants, {n_gpt} GPT-suggested")
                # Show key map summary
                km = data.get('key_map', {})
                transforms = {}
                for info in km.values():
                    t = info.get('transform', '?')
                    transforms[t] = transforms.get(t, 0) + 1
                if transforms:
                    parts = [f"{v}x {k}" for k, v in sorted(transforms.items())]
                    print(f"    Transforms: {', '.join(parts)}")
                # Show GPT suggestions summary
                gpt = data.get('gpt_suggestions', {})
                if gpt:
                    gpt_parts = []
                    for pk, info in gpt.items():
                        conf = info.get('confidence', '?')
                        gpt_parts.append(f"{pk} [{conf}]")
                    print(f"    GPT suggestions: {', '.join(gpt_parts)}")
                print()

    else:
        print(f"Unknown command: {cmd}")
        print(usage)
        sys.exit(1)


if __name__ == '__main__':
    main()
